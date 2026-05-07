
"""
Contrastive SFT Run Tracker
============================
For every run folder under --runs_root, reads:
  - trainer_state.json  → final metrics from log_history
  - adapter_config.json → all hyperparams
  - training_args.json  → all training hyperparams (if present)

Dumps everything flat into one Excel sheet, one row per run.

Usage:
    python build_experiment_tracker.py \
        --runs_root /workspace/Pupil/contrastive_experiments/outputs/contrastive_sft_v01 \
        --output    experiment_tracker.xlsx

    python /workspace/Pupil/contrastive_experiments/activity_qa_setup/run_convertor.py \
        --runs_root /workspace/Pupil/contrastive_experiments/outputs \
        --output    /workspace/Pupil/contrastive_experiments/activity_qa_setup/experiment_tracker.xlsx
"""
import argparse, json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Columns to show first, in priority order (based on your training script)
PRIORITY_COLS = [
    "run_name",
    # Contrastive-specific
    "args.negative_strategy",
    "args.contrastive_weight",
    "args.contrastive_temperature",
    "args.alpha_grounding_penalty",
    "args.num_negatives",
    "args.use_hard_negatives",
    "args.use_memory_queue",
    # Results
    "last_train.loss",
    "last_train.learning_rate",
    "last_train.epoch",
    "last_train.grad_norm",
    "last_eval.eval_loss",
    "last_eval.eval_accuracy",
    # Core training
    "args.num_train_epochs",
    "state.global_step",
    "args.per_device_train_batch_size",
    "args.gradient_accumulation_steps",
    "args.learning_rate",
    "args.vision_lr",
    "args.merger_lr",
    "args.lr_scheduler_type",
    "args.warmup_ratio",
    "args.weight_decay",
    "args.max_seq_length",
    "args.nframes",
    # LoRA
    "adapter.r",
    "adapter.lora_alpha",
    "adapter.lora_dropout",
    "adapter.target_modules",
    "adapter.base_model_name_or_path",
    # Freeze config
    "args.freeze_llm",
    "args.freeze_vision_tower",
    "args.freeze_merger",
    "args.unfreeze_topk_llm",
    "args.unfreeze_topk_vision",
]

# Columns to drop entirely (noise)
DROP_COLS = {
    "adapter.auto_mapping.base_model_class",
    "adapter.auto_mapping.parent_library",
    "adapter.corda_config",
    "adapter.eva_config",
    "adapter.exclude_modules",
    "adapter.fan_in_fan_out",
    "adapter.layer_replication",
    "adapter.layers_pattern",
    "adapter.layers_to_transform",
    "adapter.lora_bias",
    "adapter.megatron_config",
    "adapter.megatron_core",
    "adapter.modules_to_save",
    "adapter.qalora_group_size",
    "adapter.revision",
    "adapter.target_parameters",
    "adapter.trainable_token_indices",
    "adapter.use_dora",
    "adapter.use_qalora",
    "adapter.use_rslora",
    "state.is_hyper_param_search",
    "state.is_local_process_zero",
    "state.is_world_process_zero",
    "state.num_input_tokens_seen",
    "state.stateful_callbacks",
    "state.total_flos",
    "state.trial_name",
    "state.trial_params",
    "state.train_batch_size",
    "adapter.inference_mode",
    "adapter.init_lora_weights",
}


def flat(d, prefix=""):
    out = {}
    for k, v in d.items():
        key = f"{prefix}{k}" if prefix else k
        if isinstance(v, dict):
            out.update(flat(v, key + "."))
        else:
            out[key] = v
    return out


def load_json(path):
    try:
        with open(path) as f:
            return json.load(f)
    except Exception:
        return {}


def scan_run(run_dir: Path):
    row = {"run_name": run_dir.name}

    ac = load_json(run_dir / "adapter_config.json")
    row.update({f"adapter.{k}": v for k, v in flat(ac).items()})

    # all args saved by your training script (model_args + data_args + contrastive_args + training_args)
    ca = load_json(run_dir / "custom_all_args.json")
    row.update({f"args.{k}": v for k, v in flat(ca).items()})

    ts = load_json(run_dir / "trainer_state.json")
    for k, v in ts.items():
        if k != "log_history":
            row[f"state.{k}"] = v

    log_history = ts.get("log_history", [])
    train_logs = [e for e in log_history if "loss" in e and "eval_loss" not in e]
    eval_logs  = [e for e in log_history if "eval_loss" in e]

    if train_logs:
        for k, v in train_logs[-1].items():
            row[f"last_train.{k}"] = v
    if eval_logs:
        for k, v in eval_logs[-1].items():
            row[f"last_eval.{k}"] = v

    return row


def ordered_columns(rows):
    all_keys = set()
    for row in rows:
        all_keys.update(row.keys())

    # Start with priority cols that actually exist in data
    ordered = [c for c in PRIORITY_COLS if c in all_keys]
    seen = set(ordered)

    # Append remaining (minus dropped), sorted
    rest = sorted(k for k in all_keys if k not in seen and k not in DROP_COLS)
    ordered.extend(rest)

    return ordered


def build_excel(rows, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Runs"

    if not rows:
        ws["A1"] = "No runs found."
        wb.save(output_path)
        return

    cols = ordered_columns(rows)

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell_font = Font(name="Arial", size=9)
    alt_fill = PatternFill("solid", fgColor="F0F4FA")
    white_fill = PatternFill("solid", fgColor="FFFFFF")

    # Two-tone header: priority cols get darker shade
    priority_set = set(PRIORITY_COLS)
    header_fill_pri  = PatternFill("solid", fgColor="1F3864")
    header_fill_rest = PatternFill("solid", fgColor="5B7DB1")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=9)

    for col_idx, key in enumerate(cols, start=1):
        c = ws.cell(row=1, column=col_idx, value=key)
        c.fill = header_fill_pri if key in priority_set else header_fill_rest
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
        ws.column_dimensions[c.column_letter].width = min(max(len(key) + 2, 12), 40)
    ws.row_dimensions[1].height = 36

    for row_idx, row in enumerate(rows, start=2):
        fill = alt_fill if row_idx % 2 == 0 else white_fill
        for col_idx, key in enumerate(cols, start=1):
            val = row.get(key, None)
            if isinstance(val, (list, dict)):
                val = json.dumps(val)
            c = ws.cell(row=row_idx, column=col_idx, value=val)
            c.fill = fill
            c.font = cell_font
            c.alignment = Alignment(vertical="center", wrap_text=False)
            c.border = border
        ws.row_dimensions[row_idx].height = 16

    ws.freeze_panes = "B2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(output_path)
    print(f"Saved: {output_path}  ({len(rows)} runs, {len(cols)} columns)")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--runs_root", required=True)
    parser.add_argument("--output", default="experiment_tracker.xlsx")
    args = parser.parse_args()

    root = Path(args.runs_root)
    run_dirs = sorted([d for d in root.iterdir() if d.is_dir()])
    print(f"Found {len(run_dirs)} run(s) under {root}")

    rows = []
    for d in run_dirs:
        print(f"  {d.name}")
        rows.append(scan_run(d))

    build_excel(rows, args.output)


if __name__ == "__main__":
    main()