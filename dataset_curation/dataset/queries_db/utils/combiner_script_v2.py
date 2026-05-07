import os
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# ============= CONFIGURATION =============
BASE_PATH = "/home/Pupil/dataset_curation/dataset/queries_db/final_1k/"
OUTPUT_EXCEL = "/home/Pupil/dataset_curation/dataset/queries_db/final_1k/final_1k_dataset_queries.xlsx"
SOF_FOLDERS = ["sof_audio", "sof_visual", "sof_time", "sof_priority"]
# =========================================

def extract_video_name_from_query_id(query_id):
    """
    Extract video name from query_id by removing the suffix pattern.
    Example: 'admission_of_a_partner__goodwill_valuation_part_1_clean_visual_001' 
             -> 'admission_of_a_partner__goodwill_valuation_part_1_clean'
    """
    # Remove the last part which is like '_visual_001', '_audio_002', etc.
    parts = query_id.rsplit('_', 2)  # Split from right, max 2 splits
    if len(parts) >= 3:
        return parts[0]
    return query_id

def load_metadata(metadata_file):
    """Load metadata from JSONL file and return as a dictionary keyed by linked_query_id"""
    metadata_dict = {}
    try:
        with open(metadata_file, 'r', encoding='utf-8') as f:
            for line in f:
                if line.strip():
                    data = json.loads(line)
                    query_id = data.get('linked_query_id')
                    if query_id:
                        metadata_dict[query_id] = data
    except Exception as e:
        print(f"Error loading metadata file {metadata_file}: {e}")
    return metadata_dict

def process_dataset():
    """Process all dataset files and return a list of records"""
    all_records = []
    
    for sof_folder in SOF_FOLDERS:
        folder_path = os.path.join(BASE_PATH, sof_folder)
        
        # Check if folder exists
        if not os.path.exists(folder_path):
            print(f"Folder not found: {folder_path}")
            continue
        
        # Get all files in the folder
        files = os.listdir(folder_path)
        
        # Find all query files
        query_files = [f for f in files if f.endswith('_queries.json')]
        
        if not query_files:
            print(f"No query files found in {sof_folder}")
            continue
        
        print(f"Processing {len(query_files)} files in {sof_folder}...")
        
        for query_file in query_files:
            # Construct file paths
            query_path = os.path.join(folder_path, query_file)
            metadata_file = query_file.replace('_queries.json', '_metadata.jsonl')
            metadata_path = os.path.join(folder_path, metadata_file)
            
            # Load queries
            try:
                with open(query_path, 'r', encoding='utf-8') as f:
                    queries_data = json.load(f)
            except Exception as e:
                print(f"Error loading {query_file}: {e}")
                continue
            
            # Load metadata if exists
            metadata_dict = {}
            if os.path.exists(metadata_path):
                metadata_dict = load_metadata(metadata_path)
            else:
                print(f"Warning: Metadata file not found for {query_file}")
            
            # Process queries
            for video_path, queries in queries_data.items():
                for query in queries:
                    query_id = query.get('query_id', '')
                    video_name = extract_video_name_from_query_id(query_id)
                    question = query.get('question', '')
                    ground_truth = query.get('ground_truth', '')
                    
                    # Extract annotations
                    annotations = query.get('annotations', {})
                    pipeline_mode = annotations.get('pipeline_mode', '')
                    category = annotations.get('cognitive_category', '')
                    
                    # Extract timestamp from metadata
                    timestamp = ''
                    if query_id in metadata_dict:
                        metadata = metadata_dict[query_id]
                        retrieved_segments = metadata.get('retrieved_segments', [])
                        if retrieved_segments and len(retrieved_segments) > 0:
                            start = retrieved_segments[0].get('start', '')
                            end = retrieved_segments[0].get('end', '')
                            timestamp = f"{start} - {end}"
                    
                    # Create record
                    record = {
                        'query_id': query_id,
                        'video_name': video_name,
                        'pipeline_mode': pipeline_mode,
                        'category': category,
                        'timestamp': timestamp,
                        'question': question,
                        'ground_truth': ground_truth
                    }
                    
                    all_records.append(record)
    
    return all_records

def create_excel(records, output_path):
    """Create an Excel file from records"""
    # Create DataFrame
    df = pd.DataFrame(records)
    
    # Create Excel file
    wb = Workbook()
    ws = wb.active
    ws.title = "Dataset"
    
    # Write headers
    headers = ['query_id', 'video_name', 'pipeline_mode', 'category', 'timestamp', 'question', 'ground_truth']
    ws.append(headers)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Write data
    for record in records:
        ws.append([
            record['query_id'],
            record['video_name'],
            record['pipeline_mode'],
            record['category'],
            record['timestamp'],
            record['question'],
            record['ground_truth']
        ])
    
    # Adjust column widths
    column_widths = {
        'A': 70,  # query_id
        'B': 50,  # video_name
        'C': 15,  # pipeline_mode
        'D': 25,  # category
        'E': 20,  # timestamp
        'F': 80,  # question
        'G': 50   # ground_truth
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Save workbook
    wb.save(output_path)
    print(f"Excel file created successfully: {output_path}")
    print(f"Total records: {len(records)}")

def main():
    print("Starting dataset processing...")
    print(f"Base path: {BASE_PATH}")
    print(f"SOF folders: {SOF_FOLDERS}")
    print("-" * 50)
    
    # Process dataset
    records = process_dataset()
    
    if not records:
        print("No records found!")
        return
    
    print("-" * 50)
    print(f"Total records collected: {len(records)}")
    
    # Create Excel file
    create_excel(records, OUTPUT_EXCEL)

if __name__ == "__main__":
    main()